import os
import cv2
import face_recognition
import numpy as np
import base64
import csv
from attendance_main import face_recognition
from flask import Flask, render_template, request, jsonify
from datetime import datetime
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# Load known faces
known_face_encodings = []
known_face_names = []

for filename in os.listdir('known_faces'):
    if filename.endswith(('.jpg', '.png')):
        path = os.path.join('known_faces', filename)
        img = face_recognition.load_image_file(path)
        encoding = face_recognition.face_encodings(img)
        if encoding:
            known_face_encodings.append(encoding[0])
            known_face_names.append(os.path.splitext(filename)[0])

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/scan', methods=['POST'])
def scan():
    data = request.get_json()
    img_data = data['image'].split(',')[1]  # Remove base64 header
    img_bytes = base64.b64decode(img_data)

    # Convert to numpy array
    nparr = np.frombuffer(img_bytes, np.uint8)
    frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

    # Face recognition logic
    rgb_frame = frame[:, :, ::-1]
    face_locations = face_recognition.face_locations(rgb_frame)
    face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

    for face_encoding in face_encodings:
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.5)
        face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
        best_match_index = np.argmin(face_distances)
        if matches[best_match_index]:
            name = known_face_names[best_match_index]
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            mark_attendance_csv(name, timestamp)
            mark_attendance_excel(name, timestamp)
            return jsonify({'success': True, 'message': f"Welcome, {name}!"})

    return jsonify({'success': False, 'message': "Unknown face"})

def mark_attendance_csv(name, timestamp):
    with open('attendance.csv', 'a', newline='') as f:
        writer = csv.writer(f)
        writer.writerow([name, timestamp])

def mark_attendance_excel(name, timestamp):
    filename = 'attendance.xlsx'
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(['Name', 'Timestamp'])
    else:
        wb = load_workbook(filename)
        ws = wb.active
    ws.append([name, timestamp])
    wb.save(filename)

if __name__ == '__main__':
    app.run(debug=True)
